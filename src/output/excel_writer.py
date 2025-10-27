import pandas as pd
from typing import List
from entities.pendencia import Pendencia
from services.resumo_service import ResumoService
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter
import os


class ExcelWriter:
    """
    Responsável por escrever dados consolidados em arquivo Excel.
    """
    
    @staticmethod
    def salvar_relatorio_consolidado(pendencias_consolidadas: List[Pendencia],
                                   df_resumo: pd.DataFrame,
                                   caminho_saida: str) -> None:
        """
        Salva o relatório consolidado em arquivo Excel com formatação profissional.
        
        Args:
            pendencias_consolidadas: Lista de pendências consolidadas
            df_resumo: DataFrame com dados do resumo (pode estar vazio)
            caminho_saida: Caminho onde salvar o arquivo
            
        Raises:
            PermissionError: Se não conseguir escrever no arquivo
            Exception: Outros erros durante a escrita
        """
        try:
            # Converter pendências para DataFrame
            df_pendencias = ExcelWriter._pendencias_para_dataframe(pendencias_consolidadas)
            
            # Gerar resumo das pendências
            resumo_consolidado = ResumoService.gerar_resumo(pendencias_consolidadas)
            
            # Criar workbook temporário com xlsxwriter para suporte a PivotTable
            ExcelWriter._criar_arquivo_com_pivot(df_pendencias, resumo_consolidado, caminho_saida)
                    
        except PermissionError:
            raise PermissionError(f"Não foi possível salvar o arquivo. "
                                f"Verifique se o arquivo não está aberto em outro programa: {caminho_saida}")
        except Exception as e:
            raise Exception(f"Erro ao salvar arquivo Excel: {str(e)}")
    
    @staticmethod
    def _pendencias_para_dataframe(pendencias: List[Pendencia]) -> pd.DataFrame:
        """
        Converte lista de pendências para DataFrame pandas.
        
        Args:
            pendencias: Lista de objetos Pendencia
            
        Returns:
            pd.DataFrame: DataFrame com os dados das pendências
        """
        if not pendencias:
            return pd.DataFrame()
        
        # Converter cada pendência para dicionário
        dados = [pendencia.to_dict() for pendencia in pendencias]
        
        # Criar DataFrame
        df = pd.DataFrame(dados)
        
        # Garantir ordem das colunas (mesma do arquivo original)
        colunas_ordenadas = [
            'STATUS',
            'UNIDADE_NEGOCIO', 
            'EMPRESA',
            'NOME_BANCO',
            'NOME_CONTA',
            'DATA_EXTRATO',
            'NUMERO_CONTA',
            'INFORMACAO_ADICIONAL',
            'NUMERO_EXTRATO',
            'TIPO_TRANSACAO',
            'VALOR',
            'Responsável',
            'Observação',
            'Departamento', 
            'Vencimento'
        ]
        
        # Reordenar colunas, mantendo apenas as que existem
        colunas_existentes = [col for col in colunas_ordenadas if col in df.columns]
        df = df[colunas_existentes]
        
        return df
    
    @staticmethod
    def _resumo_para_dataframe(resumo_consolidado) -> pd.DataFrame:
        """
        Converte o resumo consolidado para DataFrame pandas.
        
        Args:
            resumo_consolidado: Objeto ResumoConsolidado
            
        Returns:
            pd.DataFrame: DataFrame com os dados do resumo
        """
        # Usar o método do ResumoService para gerar dados estruturados
        dados_excel = ResumoService.gerar_dados_excel(resumo_consolidado)
        
        # Criar DataFrame
        df = pd.DataFrame(dados_excel)
        
        return df
    
    @staticmethod
    def _criar_aba_pendencias(wb: Workbook, df_pendencias: pd.DataFrame) -> None:
        """
        Cria e formata a aba de Pendências como uma tabela profissional.
        """
        ws = wb.create_sheet("Pendências", 0)  # Primeira aba
        
        # Adicionar dados do DataFrame
        for r in dataframe_to_rows(df_pendencias, index=False, header=True):
            ws.append(r)
        
        # Aplicar formatação
        ExcelWriter._formatar_aba_como_tabela(ws, df_pendencias, "Pendências")
        
        # Ajustar larguras específicas para pendências
        larguras_pendencias = {
            'A': 20,  # STATUS
            'B': 25,  # UNIDADE_NEGOCIO
            'C': 30,  # EMPRESA
            'D': 20,  # NOME_BANCO
            'E': 20,  # NOME_CONTA
            'F': 15,  # DATA_EXTRATO
            'G': 15,  # NUMERO_CONTA
            'H': 30,  # INFORMACAO_ADICIONAL
            'I': 20,  # NUMERO_EXTRATO
            'J': 15,  # TIPO_TRANSACAO
            'K': 15,  # VALOR
            'L': 20,  # Responsável
            'M': 25,  # Observação
            'N': 20,  # Departamento
            'O': 12   # Vencimento
        }
        
        for col, largura in larguras_pendencias.items():
            ws.column_dimensions[col].width = largura
    
    @staticmethod
    def _criar_arquivo_com_pivot(df_pendencias: pd.DataFrame, resumo_consolidado, caminho_saida: str) -> None:
        """
        Cria o arquivo Excel completo usando pandas + openpyxl para criar tabela dinâmica atualizável.
        Conforme especificação do README_Resumo_Pivot.md
        
        A aba Resumo contém uma fórmula PIVOT do Excel que é atualizável automaticamente.
        """
        # Usar pandas ExcelWriter com openpyxl engine
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            # 1. Escrever aba de Pendências
            df_pendencias.to_excel(writer, sheet_name='Pendências', index=False)
            
            # Obter workbook e worksheet para formatação
            workbook = writer.book
            ws_pendencias = writer.sheets['Pendências']
            
            # Forçar recálculo completo das fórmulas ao abrir o arquivo
            workbook.calculation.calcMode = 'auto'
            workbook.calculation.fullCalcOnLoad = True
            
            # Configurar aba Pendências sem estilo de tabela
            max_row = len(df_pendencias) + 1
            max_col = len(df_pendencias.columns)
            
            # Ajustar larguras das colunas Pendências
            larguras_pendencias = {
                'A': 20, 'B': 25, 'C': 30, 'D': 20, 'E': 20,
                'F': 15, 'G': 15, 'H': 30, 'I': 20, 'J': 15,
                'K': 15, 'L': 20, 'M': 25, 'N': 20, 'O': 12
            }
            for col, width in larguras_pendencias.items():
                ws_pendencias.column_dimensions[col].width = width
            
            # Definir bordas cinza escuro
            border_style = Border(
                left=Side(style='thin', color='808080'),
                right=Side(style='thin', color='808080'),
                top=Side(style='thin', color='808080'),
                bottom=Side(style='thin', color='808080')
            )
            
            # Formatar todas as células com bordas e fonte tamanho 10
            for row in ws_pendencias.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = border_style
                    cell.font = Font(name='Calibri', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Formatar header (primeira linha) - negrito
            for cell in ws_pendencias[1]:
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Formatar coluna F (DATA_EXTRATO) - formato DD/MM/AAAA
            for row in range(2, max_row + 1):
                cell = ws_pendencias[f'F{row}']
                cell.number_format = 'DD/MM/YYYY'
            
            # Formatar coluna K (VALOR) - formato Contábil sem símbolo
            for row in range(2, max_row + 1):
                cell = ws_pendencias[f'K{row}']
                cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 2. Criar aba Resumo com tabela calculada dinamicamente
            ws_resumo = workbook.create_sheet('Resumo')
            
            # Escrever "Dia útil" e data do dia útil anterior
            ws_resumo['A1'] = 'Dia útil'
            ws_resumo['A1'].font = Font(name='Calibri', size=11, bold=True)
            ws_resumo['B1'] = resumo_consolidado.dia_util_referencia
            ws_resumo['B1'].number_format = 'DD/MM/YYYY'
            
            # Criar headers da tabela resumo (linha 3, sem texto "Vencimento")
            headers = ['Departamento', 'D1', '>D+1', 'Total Geral']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_resumo.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Departamentos na ordem especificada
            departamentos = ['Cash', 'Contas a Pagar', 'Contas a Receber', 'Tesouraria']
            
            # Calculate data range (header is row 1, data starts row 2)
            data_start_row = 2
            data_end_row = max_row  # Already calculated as len(df_pendencias) + 1
            
            # Adicionar fórmulas SUMPRODUCT com referências absolutas de células
            for row_idx, depto in enumerate(departamentos, start=4):
                # Departamento
                ws_resumo.cell(row=row_idx, column=1, value=depto)
                ws_resumo.cell(row=row_idx, column=1).font = Font(name='Calibri', size=10)
                
                # D1 - SUMPRODUCT
                # Column positions: A=STATUS(1), N=Departamento(14), O=Vencimento(15)
                formula_d1 = f'=SUMPRODUCT((Pendências!$A${data_start_row}:$A${data_end_row}="Não Reconciliada")*(Pendências!$N${data_start_row}:$N${data_end_row}="{depto}")*(Pendências!$O${data_start_row}:$O${data_end_row}="D1"))'
                cell_d1 = ws_resumo.cell(row=row_idx, column=2, value=formula_d1)
                cell_d1.font = Font(name='Calibri', size=10)
                cell_d1.alignment = Alignment(horizontal='center')
                cell_d1.number_format = '0'
                
                # >D+1 - SUMPRODUCT
                formula_d_mais_1 = f'=SUMPRODUCT((Pendências!$A${data_start_row}:$A${data_end_row}="Não Reconciliada")*(Pendências!$N${data_start_row}:$N${data_end_row}="{depto}")*(Pendências!$O${data_start_row}:$O${data_end_row}=">D+1"))'
                cell_d_mais_1 = ws_resumo.cell(row=row_idx, column=3, value=formula_d_mais_1)
                cell_d_mais_1.font = Font(name='Calibri', size=10)
                cell_d_mais_1.alignment = Alignment(horizontal='center')
                cell_d_mais_1.number_format = '0'
                
                # Total Geral - SUM
                cell_total = ws_resumo.cell(row=row_idx, column=4, value=f'=B{row_idx}+C{row_idx}')
                cell_total.font = Font(name='Calibri', size=10)
                cell_total.alignment = Alignment(horizontal='center')
                cell_total.number_format = '0'
            
            # Linha Total Geral
            row_total = 4 + len(departamentos)
            ws_resumo.cell(row=row_total, column=1, value='Total Geral')
            ws_resumo.cell(row=row_total, column=1).font = Font(name='Calibri', size=10, bold=True)
            
            # Fórmulas de soma para totais
            cell_sum_d1 = ws_resumo.cell(row=row_total, column=2, value=f'=SUM(B4:B{row_total-1})')
            cell_sum_d1.font = Font(name='Calibri', size=10, bold=True)
            cell_sum_d1.alignment = Alignment(horizontal='center')
            cell_sum_d1.number_format = '0'
            
            cell_sum_d_mais_1 = ws_resumo.cell(row=row_total, column=3, value=f'=SUM(C4:C{row_total-1})')
            cell_sum_d_mais_1.font = Font(name='Calibri', size=10, bold=True)
            cell_sum_d_mais_1.alignment = Alignment(horizontal='center')
            cell_sum_d_mais_1.number_format = '0'
            
            cell_sum_total = ws_resumo.cell(row=row_total, column=4, value=f'=SUM(D4:D{row_total-1})')
            cell_sum_total.font = Font(name='Calibri', size=10, bold=True)
            cell_sum_total.alignment = Alignment(horizontal='center')
            cell_sum_total.number_format = '0'
            
            # Ajustar larguras das colunas
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 15
            ws_resumo.column_dimensions['C'].width = 15
            ws_resumo.column_dimensions['D'].width = 15
            
            # Congelar painéis
            ws_resumo.freeze_panes = 'A4'
    
    @staticmethod
    def _formatar_aba_como_tabela(ws, df: pd.DataFrame, nome_tabela: str) -> None:
        """
        Formata uma aba como tabela profissional do Excel.
        """
        if df.empty:
            return
        
        # Definir range da tabela
        max_row = len(df) + 1  # +1 para o header
        max_col = len(df.columns)
        table_range = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
        
        # Criar tabela
        table = Table(displayName=nome_tabela, ref=table_range)
        
        # Estilo da tabela
        style = TableStyleInfo(
            name="TableStyleMedium9",  # Azul profissional
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Adicionar tabela à worksheet
        ws.add_table(table)
        
        # Formatação do cabeçalho
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar formatação ao cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Formatação das células de dados
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        # Aplicar formatação aos dados
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Formatação específica para valores numéricos
                if isinstance(cell.value, (int, float)) and cell.column_letter in ['K']:  # Coluna VALOR
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Congelar painéis (primeira linha)
        ws.freeze_panes = 'A2'
    
    @staticmethod
    def _ajustar_larguras_automaticas(ws) -> None:
        """
        Ajusta automaticamente as larguras das colunas baseado no conteúdo.
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Definir largura com limite máximo
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def validar_caminho_saida(caminho_saida: str) -> bool:
        """
        Valida se é possível escrever no caminho especificado.
        
        Args:
            caminho_saida: Caminho do arquivo de saída
            
        Returns:
            bool: True se é possível escrever, False caso contrário
        """
        try:
            import os
            
            # Verificar se o diretório existe
            diretorio = os.path.dirname(caminho_saida)
            if diretorio and not os.path.exists(diretorio):
                return False
            
            # Tentar criar arquivo temporário para testar permissões
            with open(caminho_saida, 'w') as f:
                pass
            os.remove(caminho_saida)
            return True
            
        except (PermissionError, OSError):
            return False 